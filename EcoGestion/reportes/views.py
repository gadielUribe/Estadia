# reportes/views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Count, Q
from django.utils.dateparse import parse_date
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.conf import settings
import os

import pandas as pd
from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Image,
    Spacer,
    PageBreak,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

from plantas.models import plantaArbol
from mantenimiento.models import TareaMantenimiento
from herramientas.models import Herramienta
from productos.models import Producto, AsignacionProducto


def _format_datetime(dt):
    if not dt:
        return ""
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime("%Y-%m-%d %H:%M")


# --------- 1) Página principal de reportes ---------
@login_required
def reportes_inicio(request):
    return render(request, "reportes/inicio_reportes.html")


# --------- 2) Helper para obtener rol ---------
def _rol(user) -> str:
    r = getattr(user, "rol", None) or getattr(user, "tipo_usuario", None) or ""
    return str(r).lower()


# --------- HELPERS DE DISEÑO MEJORADOS ---------

def _get_membrete_path():
    """Obtiene la ruta del membrete/logo de forma robusta"""
    path_absoluto = os.path.join(settings.BASE_DIR, "static", "img", "upemor-logo.png")
    if os.path.exists(path_absoluto):
        return path_absoluto
    
    try:
        from django.contrib.staticfiles import finders
        path_alternativo = finders.find("img/upemor-logo.png")
        if path_alternativo:
            return path_alternativo
    except:
        pass
    
    print("ADVERTENCIA: No se encontró el archivo de membrete.")
    return None


def _add_membrete_to_story(story, width=120, height=60):
    """Agrega el membrete al PDF con manejo de errores"""
    membrete_path = _get_membrete_path()
    if membrete_path:
        try:
            img = Image(membrete_path, width=width, height=height)
            story.append(img)
            story.append(Spacer(1, 20))
        except Exception as e:
            print(f"Error al cargar membrete: {e}")
            story.append(Spacer(1, 30))


def _get_default_table_style(header_color="#6a67c8"):
    """
    Devuelve un estilo de tabla profesional con ajuste de texto
    Optimizado para evitar solapamiento
    """
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),  # Reducido para más espacio
            ("FONTSIZE", (0, 1), (-1, -1), 7),  # Reducido para más espacio
            # Filas alternas
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f4f4f8")]),
            # Padding para evitar que el texto toque los bordes
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            # Permitir que el texto se ajuste (wrap)
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]
    )


def _create_bar_chart(title_text, data_list, labels_list):
    """Crea un gráfico de barras con título integrado"""
    drawing = Drawing(450, 220)
    
    # Título
    title = String(
        225, 190,  # Posición ajustada
        title_text,
        textAnchor='middle',
        fontSize=11,
        fontName='Helvetica-Bold'
    )
    drawing.add(title)
    
    # Gráfico de Barras
    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 50
    bc.height = 130
    bc.width = 380
    
    bc.data = [data_list]
    bc.categoryAxis.categoryNames = labels_list
    bc.categoryAxis.labels.angle = 0  # Horizontal para mejor lectura
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.dy = -10
    
    bc.valueAxis.valueMin = 0
    bc.valueAxis.labels.fontSize = 7
    bc.barWidth = 8
    bc.groupSpacing = 3
    
    drawing.add(bc)
    return drawing


def _create_pie_chart(title_text, data_list, labels_list):
    """Crea un gráfico de pastel con título integrado"""
    drawing = Drawing(450, 220)

    # Título
    title = String(
        225, 190,
        title_text,
        textAnchor='middle',
        fontSize=11,
        fontName='Helvetica-Bold'
    )
    drawing.add(title)

    # Si no hay datos, regresamos solo el título
    if not data_list or sum(data_list) == 0:
        return drawing

    pie = Pie()
    pie.x = 150
    pie.y = 30
    pie.width = 150
    pie.height = 150
    pie.data = data_list
    pie.labels = labels_list
    pie.slices.fontSize = 7
    pie.slices.strokeWidth = 0.5
    pie.slices.strokeColor = colors.black

    drawing.add(pie)
    return drawing


# --------- 3) REPORTE DE ESPECIES ---------
@login_required
def reporte_especies(request):
    usuario = request.user
    rol = _rol(usuario)
    
    # Query base según rol
    if rol == "administrador":
        qs = plantaArbol.objects.all()
    elif rol == "gestor":
        plantas_ids = TareaMantenimiento.objects.filter(
            Q(usuario_responsable=usuario) | Q(modificado_por=usuario)
        ).values("planta_id")
        qs = plantaArbol.objects.filter(id_planta__in=plantas_ids).distinct()
    elif rol == "mantenimiento":
        plantas_ids = TareaMantenimiento.objects.filter(
            usuario_responsable=usuario
        ).values("planta_id")
        qs = plantaArbol.objects.filter(id_planta__in=plantas_ids).distinct()
    else:
        qs = plantaArbol.objects.none()

    # Filtros
    q = request.GET.get("q", "").strip()
    f_ini = parse_date(request.GET.get("f_ini")) if request.GET.get("f_ini") else None
    f_fin = parse_date(request.GET.get("f_fin")) if request.GET.get("f_fin") else None
    
    if q:
        qs = qs.filter(
            Q(nombre_comun__icontains=q)
            | Q(nombre_cientifico__icontains=q)
            | Q(descripcion__icontains=q)
        )
    if f_ini:
        qs = qs.filter(fecha_plantacion__gte=f_ini)
    if f_fin:
        qs = qs.filter(fecha_plantacion__lte=f_fin)

    total = qs.count()
    columnas = [
        "nombre_comun", "nombre_cientifico", "descripcion", "fecha_plantacion",
        "periodicidad_riego", "periodicidad_poda", "periodicidad_fumigacion",
    ]
    valores = list(qs.values(*columnas))

    # Datos para gráficas
    freq_por_cientifico = (
        qs.values("nombre_cientifico")
        .annotate(c=Count("id_planta"))
        .order_by("-c")[:10]
    )
    chart1_labels = [x["nombre_cientifico"] for x in freq_por_cientifico]
    chart1_data = [x["c"] for x in freq_por_cientifico]

    freq_riego = (
        qs.values("periodicidad_riego")
        .annotate(c=Count("id_planta"))
        .order_by("-c")
    )
    chart2_labels = [x["periodicidad_riego"] or "Sin dato" for x in freq_riego]
    chart2_data = [x["c"] for x in freq_riego]

    # Exportar a Excel
    if "excel" in request.GET:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        
        df = pd.DataFrame(valores)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Especies", index=False, startrow=7)
            workbook = writer.book
            sheet = writer.sheets["Especies"]
            
            # Membrete
            membrete_path = _get_membrete_path()
            if membrete_path:
                try:
                    logo = XLImage(membrete_path)
                    logo.width = 180
                    logo.height = 90
                    sheet.add_image(logo, "A1")
                except Exception:
                    pass
            
            # Encabezado
            sheet["A5"] = "Reporte General de Especies"
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)
            sheet["A6"] = f"Total de especies registradas: {total}"
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)
            
            # Ajustar ancho de columnas automáticamente
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
            
            # Hoja de gráficas
            sheet2 = workbook.create_sheet(title="Gráficas")
            sheet2["A1"] = "Nombre científico"
            sheet2["B1"] = "Frecuencia"
            for i, (lbl, val) in enumerate(zip(chart1_labels, chart1_data), start=2):
                sheet2[f"A{i}"] = lbl
                sheet2[f"B{i}"] = val
            sheet2["D1"] = "Periodicidad riego"
            sheet2["E1"] = "Frecuencia"
            for i, (lbl, val) in enumerate(zip(chart2_labels, chart2_data), start=2):
                sheet2[f"D{i}"] = lbl
                sheet2[f"E{i}"] = val
            
            # Ajustar columnas de gráficas
            sheet2.column_dimensions['A'].width = 30
            sheet2.column_dimensions['D'].width = 25
            
            # Gráfica 1
            if chart1_labels:
                bar_chart = BarChart()
                bar_chart.title = "Frecuencia por nombre científico"
                bar_chart.y_axis.title = "Cantidad"
                bar_chart.x_axis.title = "Especie"
                data_ref = Reference(sheet2, min_col=2, min_row=1, max_row=1 + len(chart1_data))
                cats_ref = Reference(sheet2, min_col=1, min_row=2, max_row=1 + len(chart1_data))
                bar_chart.add_data(data_ref, titles_from_data=True)
                bar_chart.set_categories(cats_ref)
                bar_chart.width = 20
                bar_chart.height = 12
                sheet2.add_chart(bar_chart, "A10")
            
            # Gráfica 2
            if chart2_labels:
                pie_chart = PieChart()
                pie_chart.title = "Distribución por periodicidad de riego"
                data_ref2 = Reference(sheet2, min_col=5, min_row=1, max_row=1 + len(chart2_data))
                cats_ref2 = Reference(sheet2, min_col=4, min_row=2, max_row=1 + len(chart2_data))
                pie_chart.add_data(data_ref2, titles_from_data=True)
                pie_chart.set_categories(cats_ref2)
                pie_chart.width = 20
                pie_chart.height = 12
                sheet2.add_chart(pie_chart, "J10")
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_especies.xlsx"'
        return response

    # Exportar a PDF
    if "pdf" in request.GET:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=36, leftMargin=36, rightMargin=36, bottomMargin=36
        )
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            "table-cell",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            spaceAfter=0,
            spaceBefore=0,
        )
        header_style = ParagraphStyle(
            "table-header",
            parent=styles["Heading5"],
            fontSize=9,
            leading=11,
            alignment=1,
            textColor=colors.whitesmoke,
        )
        story = []
        
        # Membrete
        _add_membrete_to_story(story, width=120, height=60)
        
        # Título
        story.append(Paragraph("Reporte General de Especies", styles["Title"]))
        story.append(Paragraph(f"Total de especies: {total}", styles["Normal"]))
        story.append(Spacer(1, 20))
        
        # Tabla de especies
        headers = [
            "Nombre común", "Nombre científico", "Fecha plantación",
            "Riego", "Poda", "Fumigación",
        ]
        data = [headers]
        for p in valores[:300]:
            data.append([
                p.get("nombre_comun", ""),
                p.get("nombre_cientifico", ""),
                str(p.get("fecha_plantacion", "") or ""),
                p.get("periodicidad_riego", ""),
                p.get("periodicidad_poda", ""),
                p.get("periodicidad_fumigacion", ""),
            ])
        
        # Anchos proporcionales optimizados
        table = Table(data, repeatRows=1, colWidths=[80, 80, 70, 35, 35, 55])
        table.setStyle(_get_default_table_style())
        story.append(table)
        
        # Nueva página para gráficas
        story.append(PageBreak())
        
        story.append(Paragraph("Análisis Gráfico", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        # Gráfica 1
        if chart1_labels:
            chart1 = _create_bar_chart(
                "Frecuencia por nombre científico (Top 10)",
                chart1_data,
                chart1_labels
            )
            story.append(chart1)
            story.append(Spacer(1, 15))
        
        # Gráfica 2
        if chart2_labels:
            chart2 = _create_pie_chart(
                "Distribución por periodicidad de riego",
                chart2_data,
                chart2_labels
            )
            story.append(chart2)
        
        doc.build(story)
        pdf_value = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_especies.pdf"'
        return response

    # Render HTML
    context = {
        "total": total, "plantas": valores, "q": q,
        "f_ini": request.GET.get("f_ini", ""), "f_fin": request.GET.get("f_fin", ""),
    }
    return render(request, "reportes/reportes_especies.html", context)


# --------- 4) REPORTE DE MANTENIMIENTO ---------
@login_required
def reporte_mantenimiento(request):
    usuario = request.user
    rol = _rol(usuario)
    
    # Query base según rol
    if rol == "administrador":
        qs = TareaMantenimiento.objects.select_related("planta", "usuario_responsable")
    elif rol == "gestor":
        qs = TareaMantenimiento.objects.select_related("planta", "usuario_responsable").filter(
            Q(usuario_responsable=usuario) | Q(modificado_por=usuario)
        )
    elif rol == "mantenimiento":
        qs = TareaMantenimiento.objects.select_related("planta", "usuario_responsable").filter(
            usuario_responsable=usuario
        )
    else:
        qs = TareaMantenimiento.objects.none()

    # Filtros
    q = request.GET.get("q", "").strip()
    f_ini = parse_date(request.GET.get("f_ini")) if request.GET.get("f_ini") else None
    f_fin = parse_date(request.GET.get("f_fin")) if request.GET.get("f_fin") else None
    
    if q:
        qs = qs.filter(
            Q(planta__nombre_comun__icontains=q)
            | Q(planta__nombre_cientifico__icontains=q)
            | Q(observaciones__icontains=q)
            | Q(tipo__icontains=q)
        )
    if f_ini:
        qs = qs.filter(fecha_programada__date__gte=f_ini)
    if f_fin:
        qs = qs.filter(fecha_programada__date__lte=f_fin)

    total_tareas = qs.count()
    realizadas = qs.filter(estado=TareaMantenimiento.ESTADO_REALIZADA).count()
    cumplimiento_global = round(realizadas * 100 / total_tareas, 2) if total_tareas > 0 else 0.0

    # Datos tabulares
    filas = []
    for t in qs:
        filas.append({
            "especie": t.planta.nombre_comun if t.planta else "",
            "tipo": t.get_tipo_display(),
            "fecha_programada": t.fecha_programada,
            "fecha_realizacion": t.fecha_realizacion,
            "estado": t.get_estado_display(),
            "observaciones": t.observaciones or "",
        })

    # Resumen por tipo de tarea
    tipos_stats = []
    chart_tipo_labels = []
    chart_tipo_data = []
    for tipo_code, tipo_label in TareaMantenimiento.TIPOS:
        total_tipo = qs.filter(tipo=tipo_code).count()
        realizadas_tipo = qs.filter(tipo=tipo_code, estado=TareaMantenimiento.ESTADO_REALIZADA).count()
        pendientes_tipo = total_tipo - realizadas_tipo
        pct_tipo = round(realizadas_tipo * 100 / total_tipo, 2) if total_tipo > 0 else 0.0
        tipos_stats.append(
            {
                "tipo_codigo": tipo_code,
                "tipo": tipo_label,
                "total": total_tipo,
                "realizadas": realizadas_tipo,
                "pendientes": pendientes_tipo,
                "porcentaje": pct_tipo,
            }
        )
        if total_tipo > 0:
            chart_tipo_labels.append(tipo_label)
            chart_tipo_data.append(pct_tipo)

    chart1_labels = chart_tipo_labels
    chart1_data = chart_tipo_data

    # Por especie (se mantiene para análisis opcional)
    por_especie = (
        qs.values("planta__nombre_comun")
        .annotate(
            total=Count("id"),
            realizadas=Count("id", filter=Q(estado=TareaMantenimiento.ESTADO_REALIZADA)),
        )
        .order_by("-total")[:10]
    )
    chart_especie_labels = []
    chart_especie_data = []
    for r in por_especie:
        nombre_e = r["planta__nombre_comun"] or "Sin especie"
        total_e = r["total"] or 0
        realizadas_e = r["realizadas"] or 0
        pct_e = round(realizadas_e * 100 / total_e, 2) if total_e > 0 else 0.0
        chart_especie_labels.append(nombre_e)
        chart_especie_data.append(pct_e)
    chart2_labels = chart_especie_labels
    chart2_data = chart_especie_data

        # Exportar a Excel
    if "excel" in request.GET:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter
        from django.utils import timezone  # por si no lo tienes ya importado arriba

        # ---- Normalizar fechas para que no tengan timezone (o sean texto) ----
        filas_excel = []
        for f in filas:
            row = f.copy()
            for campo_fecha in ("fecha_programada", "fecha_realizacion"):
                row[campo_fecha] = _format_datetime(row.get(campo_fecha))
            filas_excel.append(row)

        # Usamos filas_excel en lugar de filas
        df = pd.DataFrame(filas_excel)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Mantenimiento", index=False, startrow=7)
            workbook = writer.book
            sheet = writer.sheets["Mantenimiento"]
            
            # Membrete
            membrete_path = _get_membrete_path()
            if membrete_path:
                try:
                    logo = XLImage(membrete_path)
                    logo.width = 180
                    logo.height = 90
                    sheet.add_image(logo, "A1")
                except Exception:
                    pass
            
            # Encabezado
            sheet["A5"] = "Reporte de mantenimiento programado y realizado"
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)
            sheet["A6"] = f"Total de tareas: {total_tareas} | Cumplimiento global: {cumplimiento_global}%"
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max() if not df.empty else 0,
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
            
            # Hoja de resumen por tipo
            sheet2 = workbook.create_sheet(title="Resumen tipos")
            sheet2["A1"] = "Tipo"
            sheet2["B1"] = "Total"
            sheet2["C1"] = "Realizadas"
            sheet2["D1"] = "Pendientes"
            sheet2["E1"] = "% Cumplimiento"
            for idx, stat in enumerate(tipos_stats, start=2):
                sheet2[f"A{idx}"] = stat["tipo"]
                sheet2[f"B{idx}"] = stat["total"]
                sheet2[f"C{idx}"] = stat["realizadas"]
                sheet2[f"D{idx}"] = stat["pendientes"]
                sheet2[f"E{idx}"] = stat["porcentaje"]
            sheet2.column_dimensions["A"].width = 18
            sheet2.column_dimensions["E"].width = 15
            
            if tipos_stats:
                bar1 = BarChart()
                bar1.title = "Cumplimiento por tipo (%)"
                bar1.y_axis.title = "% Cumplimiento"
                bar1.x_axis.title = "Tipo"
                last_row = 1 + len(tipos_stats)
                data_ref1 = Reference(sheet2, min_col=5, min_row=1, max_row=last_row)
                cats_ref1 = Reference(sheet2, min_col=1, min_row=2, max_row=last_row)
                bar1.add_data(data_ref1, titles_from_data=True)
                bar1.set_categories(cats_ref1)
                bar1.width = 18
                bar1.height = 10
                sheet2.add_chart(bar1, "G2")
            
            if chart_especie_labels:
                sheet3 = workbook.create_sheet(title="Top especies")
                sheet3["A1"] = "Especie"
                sheet3["B1"] = "% Cumplimiento"
                for idx, (lbl, val) in enumerate(zip(chart_especie_labels, chart_especie_data), start=2):
                    sheet3[f"A{idx}"] = lbl
                    sheet3[f"B{idx}"] = float(val)
                sheet3.column_dimensions["A"].width = 30
                bar2 = BarChart()
                bar2.title = "Cumplimiento por especie (%)"
                bar2.y_axis.title = "% Cumplimiento"
                bar2.x_axis.title = "Especie"
                data_ref2 = Reference(sheet3, min_col=2, min_row=1, max_row=1 + len(chart_especie_data))
                cats_ref2 = Reference(sheet3, min_col=1, min_row=2, max_row=1 + len(chart_especie_data))
                bar2.add_data(data_ref2, titles_from_data=True)
                bar2.set_categories(cats_ref2)
                bar2.width = 18
                bar2.height = 10
                sheet3.add_chart(bar2, "D2")
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_mantenimiento.xlsx"'
        return response


    # Exportar a PDF
    if "pdf" in request.GET:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=36, leftMargin=36, rightMargin=36, bottomMargin=36
        )
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            "table-cell",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            spaceAfter=0,
            spaceBefore=0,
        )
        header_style = ParagraphStyle(
            "table-header",
            parent=styles["Heading5"],
            fontSize=9,
            leading=11,
            alignment=1,
            textColor=colors.whitesmoke,
        )
        story = []
        
        # Membrete
        _add_membrete_to_story(story, width=120, height=60)
        
        # Título y datos generales
        story.append(Paragraph("Reporte de mantenimiento programado y realizado", styles["Title"]))
        story.append(Paragraph(
            f"Total de tareas: {total_tareas} | Cumplimiento global: {cumplimiento_global}%",
            styles["Normal"],
        ))
        story.append(Spacer(1, 20))
        
        # Tabla de tareas - ANCHOS AJUSTADOS PARA TÍTULOS COMPLETOS
        headers = [
            "Especie", "Tipo", "Fecha programada", "Fecha realización",
            "Estado", "Observaciones",
        ]
        data = [[Paragraph(h, header_style) for h in headers]]
        for f in filas[:300]:
            data.append([
                Paragraph(f["especie"] or "-", cell_style),
                Paragraph(f["tipo"] or "-", cell_style),
                Paragraph(_format_datetime(f["fecha_programada"]), cell_style),
                Paragraph(_format_datetime(f["fecha_realizacion"]), cell_style),
                Paragraph(f["estado"] or "-", cell_style),
                Paragraph(f["observaciones"] or "-", cell_style),
            ])
        
        table = Table(data, repeatRows=1, colWidths=[65, 60, 75, 75, 55, 160])
        table.setStyle(_get_default_table_style())
        story.append(table)

        # Resumen por tipo
        story.append(Spacer(1, 18))
        story.append(Paragraph("Resumen de cumplimiento por tipo", styles["Heading3"]))
        resumen_headers = ["Tipo", "Total", "Realizadas", "Pendientes", "% Cumplimiento"]
        resumen_data = [[Paragraph(h, header_style) for h in resumen_headers]]
        for stat in tipos_stats:
            resumen_data.append(
                [
                    Paragraph(stat["tipo"], cell_style),
                    Paragraph(str(stat["total"]), cell_style),
                    Paragraph(str(stat["realizadas"]), cell_style),
                    Paragraph(str(stat["pendientes"]), cell_style),
                    Paragraph(f'{stat["porcentaje"]:.2f}', cell_style),
                ]
            )
        resumen_table = Table(resumen_data, repeatRows=1, colWidths=[90, 50, 60, 60, 70])
        resumen_table.setStyle(_get_default_table_style("#4a69bd"))
        story.append(resumen_table)
        
        # Nueva página para gráficas
        story.append(PageBreak())
        
        story.append(Paragraph("Análisis Gráfico", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        # Gráfica 1
        if chart1_labels:
            chart1 = _create_bar_chart(
                "Cumplimiento por tipo (%)",
                chart1_data,
                chart1_labels
            )
            story.append(chart1)
            story.append(Spacer(1, 15))
        
        # Gráfica 2
        if chart2_labels:
            chart2 = _create_bar_chart(
                "Cumplimiento por especie (%) (Top 10)",
                chart2_data,
                chart2_labels
            )
            story.append(chart2)
        
        doc.build(story)
        pdf_value = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_mantenimiento.pdf"'
        return response

    # Render HTML
    context = {
        "total_tareas": total_tareas,
        "realizadas": realizadas,
        "cumplimiento_global": cumplimiento_global,
        "tareas": filas,
        "tipos_stats": tipos_stats,
        "q": q,
        "f_ini": request.GET.get("f_ini", ""),
        "f_fin": request.GET.get("f_fin", ""),
    }
    return render(request, "reportes/reportes_mantenimiento.html", context)


# --------- 5) REPORTE DE CUMPLIMIENTO ---------
@login_required
def reportes_cumplimiento_usuarios(request):
    usuario = request.user
    rol = _rol(usuario)
    if rol not in ("administrador", "gestor"):
        return HttpResponseForbidden("No tiene permiso para ver este reporte.")

    User = get_user_model()
    qs = TareaMantenimiento.objects.select_related("usuario_responsable").filter(
        usuario_responsable__rol="mantenimiento"
    )

    # Filtros
    q = request.GET.get("q", "").strip()
    f_ini = parse_date(request.GET.get("f_ini")) if request.GET.get("f_ini") else None
    f_fin = parse_date(request.GET.get("f_fin")) if request.GET.get("f_fin") else None
    
    user_ids = []
    if q:
        user_ids = list(
            User.objects.filter(
                Q(nombre_completo__icontains=q)
                | Q(matricula__icontains=q)
                | Q(email__icontains=q)
            )
            .values_list("id_usuario", flat=True)
        )
    if f_ini:
        qs = qs.filter(fecha_programada__date__gte=f_ini)
    if f_fin:
        qs = qs.filter(fecha_programada__date__lte=f_fin)

    total_tareas = qs.count()
    realizadas = qs.filter(estado=TareaMantenimiento.ESTADO_REALIZADA).count()
    pendientes = qs.filter(estado=TareaMantenimiento.ESTADO_PENDIENTE).count()
    cumplimiento_global = round(realizadas * 100 / total_tareas, 2) if total_tareas > 0 else 0.0

    # Agregados por usuario
    agregados = (
        qs.values(
            "usuario_responsable__id_usuario", "usuario_responsable__nombre_completo",
            "usuario_responsable__matricula", "usuario_responsable__email",
        )
        .annotate(
            total=Count("id"),
            realizadas=Count("id", filter=Q(estado=TareaMantenimiento.ESTADO_REALIZADA)),
        )
        .order_by("-realizadas")
    )

    filas = []
    filtered_filas = []
    for r in agregados:
        total_u = r["total"] or 0
        realizadas_u = r["realizadas"] or 0
        pendientes_u = total_u - realizadas_u
        pct = round(realizadas_u * 100 / total_u, 2) if total_u > 0 else 0.0
        
        nombre = (
            r["usuario_responsable__nombre_completo"]
            or r["usuario_responsable__matricula"]
            or r["usuario_responsable__email"]
            or "Sin nombre"
        )
        include = True
        if q and user_ids:
            include = r["usuario_responsable__id_usuario"] in user_ids
        filas.append({
            "id": r["usuario_responsable__id_usuario"], "nombre": nombre,
            "matricula": r["usuario_responsable__matricula"],
            "email": r["usuario_responsable__email"],
            "total": total_u, "realizadas": realizadas_u,
            "pendientes": pendientes_u, "porcentaje": pct,
        })
        if include:
            filtered_filas.append(filas[-1])
    filas = filtered_filas if q and user_ids else filas

    # Datos para gráficas
    chart1_labels = [f["nombre"] for f in filas]
    chart1_data = [f["porcentaje"] for f in filas]
    chart2_labels = ["Realizadas", "Pendientes"]
    chart2_data = [realizadas, pendientes]

    # Exportar a Excel
    if "excel" in request.GET:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
        
        filas_excel = []
        for fila in filas:
            row = fila.copy()
            fh = row.get("fecha_hora")
            if fh:
                if timezone.is_aware(fh):
                    fh = timezone.localtime(fh)
                fh = fh.strftime("%Y-%m-%d %H:%M")
            else:
                fh = ""
            row["fecha_hora"] = fh
            filas_excel.append(row)

        df = pd.DataFrame(filas_excel)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Cumplimiento", index=False, startrow=7)
            workbook = writer.book
            sheet = writer.sheets["Cumplimiento"]
            
            # Membrete
            membrete_path = _get_membrete_path()
            if membrete_path:
                try:
                    logo = XLImage(membrete_path)
                    logo.width = 180
                    logo.height = 90
                    sheet.add_image(logo, "A1")
                except Exception:
                    pass
            
            # Encabezado
            sheet["A5"] = "Reporte de cumplimiento de tareas por usuarios"
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)
            sheet["A6"] = f"Total tareas: {total_tareas} | Realizadas: {realizadas} | Pendientes: {pendientes} | Cumplimiento global: {cumplimiento_global}%"
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max() if not df.empty else 0,
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
            
            # Hoja de gráficas
            sheet2 = workbook.create_sheet(title="Gráficas")
            sheet2["A1"] = "Usuario"
            sheet2["B1"] = "% Cumplimiento"
            for i, (lbl, val) in enumerate(zip(chart1_labels, chart1_data), start=2):
                sheet2[f"A{i}"] = lbl
                sheet2[f"B{i}"] = float(val)
            
            sheet2["D1"] = "Estado"
            sheet2["E1"] = "Cantidad"
            for i, (lbl, val) in enumerate(zip(chart2_labels, chart2_data), start=2):
                sheet2[f"D{i}"] = lbl
                sheet2[f"E{i}"] = int(val)
            
            # Ajustar columnas de gráficas
            sheet2.column_dimensions['A'].width = 30
            sheet2.column_dimensions['D'].width = 15
            
            # Gráfica 1
            if chart1_labels:
                bar1 = BarChart()
                bar1.title = "Cumplimiento por usuario (%)"
                bar1.y_axis.title = "% Cumplimiento"
                bar1.x_axis.title = "Usuario"
                data_ref1 = Reference(sheet2, min_col=2, min_row=1, max_row=1 + len(chart1_data))
                cats_ref1 = Reference(sheet2, min_col=1, min_row=2, max_row=1 + len(chart1_data))
                bar1.add_data(data_ref1, titles_from_data=True)
                bar1.set_categories(cats_ref1)
                bar1.width = 20
                bar1.height = 12
                sheet2.add_chart(bar1, "A10")
            
            # Gráfica 2
            if sum(chart2_data) > 0:
                pie = PieChart()
                pie.title = "Tareas realizadas vs pendientes"
                data_ref2 = Reference(sheet2, min_col=5, min_row=1, max_row=1 + len(chart2_data))
                cats_ref2 = Reference(sheet2, min_col=4, min_row=2, max_row=1 + len(chart2_data))
                pie.add_data(data_ref2, titles_from_data=True)
                pie.set_categories(cats_ref2)
                pie.width = 15
                pie.height = 10
                sheet2.add_chart(pie, "J10")
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_cumplimiento_usuarios.xlsx"'
        return response

    # Exportar a PDF
    if "pdf" in request.GET:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=36, leftMargin=36, rightMargin=36, bottomMargin=36
        )
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            "table-cell",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            spaceAfter=0,
            spaceBefore=0,
        )
        header_style = ParagraphStyle(
            "table-header",
            parent=styles["Heading5"],
            fontSize=9,
            leading=11,
            alignment=1,
            textColor=colors.whitesmoke,
        )
        story = []
        
        # Membrete
        _add_membrete_to_story(story, width=120, height=60)
        
        # Título y datos generales
        story.append(Paragraph("Reporte de cumplimiento de tareas por usuarios", styles["Title"]))
        story.append(Paragraph(
            f"Total tareas: {total_tareas} | Realizadas: {realizadas} | Pendientes: {pendientes} | Cumplimiento global: {cumplimiento_global}%",
            styles["Normal"],
        ))
        story.append(Spacer(1, 20))
        
        # Tabla de usuarios - ANCHOS AJUSTADOS PARA TÍTULOS COMPLETOS
        headers = [
            "Usuario", "Matrícula", "Email",
            "Tareas programadas", "Realizadas", "Pendientes", "% Cumplimiento",
        ]
        data = [[Paragraph(h, header_style) for h in headers]]
        for f in filas[:300]:
            data.append([
                Paragraph(f["nombre"] or "-", cell_style),
                Paragraph(f["matricula"] or "-", cell_style),
                Paragraph(f["email"] or "-", cell_style),
                Paragraph(str(f["total"]), cell_style),
                Paragraph(str(f["realizadas"]), cell_style),
                Paragraph(str(f["pendientes"]), cell_style),
                Paragraph(f'{f["porcentaje"]:.2f}', cell_style),
            ])
        
        # ANCHOS OPTIMIZADOS: más espacio para "Tareas programadas" y "Email"
        table = Table(data, repeatRows=1, colWidths=[85, 60, 120, 70, 55, 55, 60])
        table.setStyle(_get_default_table_style())
        story.append(table)
        
        # Nueva página para gráficas
        story.append(PageBreak())
        
        story.append(Paragraph("Análisis Gráfico", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        # Gráfica 1
        if chart1_labels:
            chart1 = _create_bar_chart(
                "Cumplimiento por usuario (%)",
                chart1_data,
                chart1_labels
            )
            story.append(chart1)
            story.append(Spacer(1, 15))
        
        # Gráfica 2
        if sum(chart2_data) > 0:
            chart2 = _create_pie_chart(
                "Tareas realizadas vs pendientes",
                chart2_data,
                chart2_labels
            )
            story.append(chart2)
        
        doc.build(story)
        pdf_value = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_cumplimiento_usuarios.pdf"'
        return response

    # Render HTML
    context = {
        "total_tareas": total_tareas, "realizadas": realizadas, "pendientes": pendientes,
        "cumplimiento_global": cumplimiento_global, "usuarios": filas, "q": q,
        "f_ini": request.GET.get("f_ini", ""), "f_fin": request.GET.get("f_fin", ""),
        "chart1_labels": chart1_labels, "chart1_data": chart1_data,
        "chart2_labels": chart2_labels, "chart2_data": chart2_data,
    }
    return render(request, "reportes/reportes_tareas.html", context)


# --------- 6) REPORTE DE ACTIVIDADES ---------
@login_required
def reporte_actividades_usuario(request):
    usuario = request.user
    rol = _rol(usuario)
    if rol not in ("administrador", "gestor"):
        return HttpResponseForbidden("No tiene permiso para ver este reporte.")

    User = get_user_model()
    usuarios_mantenimiento = User.objects.filter(
        rol="mantenimiento", is_active=True
    ).order_by("nombre_completo")

    usuario_id = request.GET.get("usuario")
    tipo_accion = request.GET.get("tipo", "todas")
    f_ini = parse_date(request.GET.get("f_ini")) if request.GET.get("f_ini") else None
    f_fin = parse_date(request.GET.get("f_fin")) if request.GET.get("f_fin") else None

    # Query base
    qs = TareaMantenimiento.objects.select_related("usuario_responsable").filter(
        usuario_responsable__rol="mantenimiento"
    )
    if usuario_id:
        qs = qs.filter(usuario_responsable__id_usuario=usuario_id)
    else:
        qs = qs.none()
    if tipo_accion == "realizada":
        qs = qs.filter(estado=TareaMantenimiento.ESTADO_REALIZADA)
    elif tipo_accion == "pendiente":
        qs = qs.filter(estado=TareaMantenimiento.ESTADO_PENDIENTE)
    if f_ini:
        qs = qs.filter(fecha_programada__date__gte=f_ini)
    if f_fin:
        qs = qs.filter(fecha_programada__date__lte=f_fin)

    # Datos tabulares
    filas = []
    for t in qs:
        u = t.usuario_responsable
        if not u:
            nombre_usuario = "Sin usuario"
            rol_usuario = "-"
            matricula = ""
            email = ""
        else:
            nombre_usuario = u.nombre_completo or u.matricula or u.email
            rol_usuario = u.get_rol_display() if hasattr(u, "get_rol_display") else u.rol
            matricula = u.matricula
            email = u.email
        
        tipo = (
            "Tarea de mantenimiento - Realizada"
            if t.estado == TareaMantenimiento.ESTADO_REALIZADA
            else "Tarea de mantenimiento - Pendiente"
        )
        fecha_hora = t.fecha_realizacion or t.fecha_programada if t.estado == TareaMantenimiento.ESTADO_REALIZADA else t.fecha_programada
        detalle = f"{t.get_tipo_display()} en {t.planta.nombre_comun if t.planta else 'Sin especie'}"
        
        filas.append({
            "usuario": nombre_usuario, "rol": rol_usuario, "matricula": matricula,
            "email": email, "tipo_accion": tipo, "fecha_hora": fecha_hora,
            "detalle": detalle,
        })

    # Estadísticas
    total_acciones = len(filas)
    realizadas = qs.filter(estado=TareaMantenimiento.ESTADO_REALIZADA).count()
    pendientes = qs.filter(estado=TareaMantenimiento.ESTADO_PENDIENTE).count()
    cumplimiento = round(realizadas * 100 / total_acciones, 2) if total_acciones > 0 else 0.0

    tipo_counts = {"riego": 0, "poda": 0, "fumigacion": 0}
    if usuario_id:
        tipo_counts["riego"] = qs.filter(tipo=TareaMantenimiento.TIPO_RIEGO, estado=TareaMantenimiento.ESTADO_REALIZADA).count()
        tipo_counts["poda"] = qs.filter(tipo=TareaMantenimiento.TIPO_PODA, estado=TareaMantenimiento.ESTADO_REALIZADA).count()
        tipo_counts["fumigacion"] = qs.filter(tipo=TareaMantenimiento.TIPO_FUMIGACION, estado=TareaMantenimiento.ESTADO_REALIZADA).count()

    # Exportar a Excel
    if "excel" in request.GET:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter
        
        filas_excel = []
        for fila in filas:
            row = fila.copy()
            row["fecha_hora"] = _format_datetime(row.get("fecha_hora"))
            filas_excel.append(row)

        df = pd.DataFrame(filas_excel)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Actividades", index=False, startrow=7)
            workbook = writer.book
            sheet = writer.sheets["Actividades"]
            
            # Membrete
            membrete_path = _get_membrete_path()
            if membrete_path:
                try:
                    logo = XLImage(membrete_path)
                    logo.width = 180
                    logo.height = 90
                    sheet.add_image(logo, "A1")
                except Exception:
                    pass
            
            # Encabezado
            sheet["A5"] = "Reporte de actividades de usuario"
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)
            sheet["A6"] = (
                f"Total acciones: {total_acciones} | "
                f"Realizadas: {realizadas} | Pendientes: {pendientes} | "
                f"% cumplimiento (tareas realizadas): {cumplimiento}%"
            )
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max() if not df.empty else 0,
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
            
            # Hoja de gráficas
            if usuario_id:
                sheet2 = workbook.create_sheet(title="Gráficas")
                sheet2["A1"] = "Tipo"
                sheet2["B1"] = "Realizadas"
                sheet2["A2"] = "Riego"
                sheet2["B2"] = tipo_counts["riego"]
                sheet2["A3"] = "Poda"
                sheet2["B3"] = tipo_counts["poda"]
                sheet2["A4"] = "Fumigación"
                sheet2["B4"] = tipo_counts["fumigacion"]
                sheet2.column_dimensions["A"].width = 20
                chart = BarChart()
                chart.title = "Tareas realizadas por tipo"
                chart.y_axis.title = "Cantidad"
                chart.x_axis.title = "Tipo"
                data_ref = Reference(sheet2, min_col=2, min_row=1, max_row=4)
                cats_ref = Reference(sheet2, min_col=1, min_row=2, max_row=4)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 18
                chart.height = 10
                sheet2.add_chart(chart, "D2")
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_actividades_usuario.xlsx"'
        return response

    # Exportar a PDF
    if "pdf" in request.GET:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=36, leftMargin=36, rightMargin=36, bottomMargin=36
        )
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            "table-cell",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            spaceBefore=0,
            spaceAfter=0,
        )
        header_style = ParagraphStyle(
            "table-header",
            parent=styles["Heading5"],
            fontSize=9,
            leading=11,
            alignment=1,
            textColor=colors.whitesmoke,
        )
        story = []
        
        # Membrete
        _add_membrete_to_story(story, width=120, height=60)
        
        # Título y estadísticas
        story.append(Paragraph("Reporte de actividades de usuario", styles["Title"]))
        story.append(Paragraph(
            f"Total acciones: {total_acciones} | Realizadas: {realizadas} | Pendientes: {pendientes} | % cumplimiento: {cumplimiento}%",
            styles["Normal"],
        ))
        story.append(Spacer(1, 20))
        
        # Tabla de actividades - ANCHOS AJUSTADOS PARA TÍTULOS COMPLETOS
        headers = ["Usuario", "Rol", "Tipo de acción", "Fecha y hora", "Detalle"]
        data = [[Paragraph(h, header_style) for h in headers]]
        for f in filas[:500]:
            data.append([
                Paragraph(f["usuario"] or "-", cell_style),
                Paragraph(f["rol"] or "-", cell_style),
                Paragraph(f["tipo_accion"] or "-", cell_style),
                Paragraph(_format_datetime(f["fecha_hora"]), cell_style),
                Paragraph(f["detalle"] or "-", cell_style),
            ])
        
        # ANCHOS OPTIMIZADOS: más espacio para "Detalle" y "Tipo de acción"
        table = Table(data, repeatRows=1, colWidths=[75, 45, 90, 70, 220])
        table.setStyle(_get_default_table_style())
        story.append(table)
        
        # Nueva página para gráficas
        story.append(PageBreak())
        
        story.append(Paragraph("Análisis Gráfico", styles["Heading2"]))
        story.append(Spacer(1, 12))
        
        if usuario_id:
            chart1 = _create_bar_chart(
                "Tareas realizadas por tipo",
                [tipo_counts["riego"], tipo_counts["poda"], tipo_counts["fumigacion"]],
                ["Riego", "Poda", "Fumigación"],
            )
            story.append(chart1)
        
        doc.build(story)
        pdf_value = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_actividades_usuario.pdf"'
        return response

    # Render HTML
    context = {
        "usuarios_mantenimiento": usuarios_mantenimiento,
        "usuario_id": usuario_id or "",
        "tipo_accion": tipo_accion,
        "f_ini": request.GET.get("f_ini", ""),
        "f_fin": request.GET.get("f_fin", ""),
        "total_acciones": total_acciones,
        "realizadas": realizadas,
        "pendientes": pendientes,
        "cumplimiento": cumplimiento,
        "actividades": filas,
        "tipo_counts": tipo_counts,
    }
    return render(request, "reportes/reportes_actividades_usuario.html", context)


# --------- 7) REPORTE USO HERRAMIENTAS ---------
@login_required
def reporte_uso_herramientas(request):
    usuario = request.user
    rol = _rol(usuario)
    if rol not in ("administrador", "gestor", "mantenimiento"):
        return HttpResponseForbidden("No tiene permiso para ver este reporte.")

    # Filtros
    q = request.GET.get("q", "").strip()
    estado_filtro = request.GET.get("estado", "").strip()
    f_ini = parse_date(request.GET.get("f_ini")) if request.GET.get("f_ini") else None
    f_fin = parse_date(request.GET.get("f_fin")) if request.GET.get("f_fin") else None

    # Query de herramientas
    herramientas_qs = Herramienta.objects.all()
    if q:
        herramientas_qs = herramientas_qs.filter(
            Q(nombre__icontains=q) | Q(descripcion__icontains=q)
        )
    if estado_filtro:
        herramientas_qs = herramientas_qs.filter(estado=estado_filtro)

    herramientas = list(herramientas_qs)
    herramientas_ids = [h.id_herramienta for h in herramientas]
    total_herramientas = len(herramientas)

    # Tareas asociadas
    tareas_qs = TareaMantenimiento.objects.filter(herramienta_id__in=herramientas_ids).select_related(
        "planta", "usuario_responsable"
    )
    if f_ini:
        tareas_qs = tareas_qs.filter(fecha_programada__date__gte=f_ini)
    if f_fin:
        tareas_qs = tareas_qs.filter(fecha_programada__date__lte=f_fin)

    # Estadísticas por herramienta
    stats = {hid: {"riego": 0, "poda": 0, "fumigacion": 0, "total": 0} for hid in herramientas_ids}
    for t in tareas_qs:
        if not t.herramienta_id:
            continue
        s = stats.setdefault(
            t.herramienta_id, {"riego": 0, "poda": 0, "fumigacion": 0, "total": 0}
        )
        s["total"] += 1
        if t.tipo == TareaMantenimiento.TIPO_RIEGO:
            s["riego"] += 1
        elif t.tipo == TareaMantenimiento.TIPO_PODA:
            s["poda"] += 1
        elif t.tipo == TareaMantenimiento.TIPO_FUMIGACION:
            s["fumigacion"] += 1

    # Filas para tabla principal
    filas = []
    for h in herramientas:
        s = stats.get(h.id_herramienta, {"riego": 0, "poda": 0, "fumigacion": 0, "total": 0})
        filas.append({
            "id": h.id_herramienta, "nombre": h.nombre, "descripcion": h.descripcion,
            "estado": h.get_estado_display() if hasattr(h, "get_estado_display") else h.estado,
            "total": s["total"], "riego": s["riego"], "poda": s["poda"], "fumigacion": s["fumigacion"],
        })

    # Historial (solo admin y gestor)
    historial = []
    if rol in ("administrador", "gestor"):
        for t in tareas_qs.order_by("-fecha_programada")[:500]:
            historial.append({
                "herramienta": t.herramienta.nombre if t.herramienta else "",
                "tipo": t.get_tipo_display(),
                "planta": t.planta.nombre_comun if t.planta else "",
                "fecha_programada": t.fecha_programada,
                "estado": t.get_estado_display(),
                "usuario": t.usuario_responsable.nombre_completo if t.usuario_responsable else "",
            })

    # Exportar a Excel
    if "excel" in request.GET:
        if rol == "mantenimiento":
            return HttpResponseForbidden("Solo puede consultar disponibilidad, no exportar reportes.")
        
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        
        df = pd.DataFrame(filas)
        # Normalizar fechas de historial a texto sin TZ
        def _normalize_hist(row):
            row = row.copy()
            for campo in ("fecha_programada",):
                val = row.get(campo)
                row[campo] = _format_datetime(val)
            return row

        df_hist = pd.DataFrame([_normalize_hist(h) for h in historial]) if historial else pd.DataFrame()
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja 1: Herramientas
            df.to_excel(writer, sheet_name="Herramientas", index=False, startrow=7)
            workbook = writer.book
            sheet = writer.sheets["Herramientas"]
            
            # Membrete
            membrete_path = _get_membrete_path()
            if membrete_path:
                try:
                    logo = XLImage(membrete_path)
                    logo.width = 180
                    logo.height = 90
                    sheet.add_image(logo, "A1")
                except Exception:
                    pass
            
            # Encabezado
            sheet["A5"] = "Reporte de uso y disponibilidad de herramientas"
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)
            sheet["A6"] = f"Número total de herramientas registradas: {total_herramientas}"
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max() if not df.empty else 0,
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
            
            # Hoja 2: Historial
            if not df_hist.empty:
                df_hist.to_excel(writer, sheet_name="Historial", index=False, startrow=4)
                sheet2 = writer.sheets["Historial"]
                sheet2["A1"] = "Historial de tareas vinculadas a cada herramienta"
                sheet2["A1"].font = sheet2["A1"].font.copy(bold=True, size=14)
                # Ajustar columnas del historial
                for idx, col in enumerate(df_hist.columns, 1):
                    max_length = max(
                        df_hist[col].astype(str).apply(len).max() if not df_hist.empty else 0,
                        len(col)
                    )
                    adjusted_width = min(max_length + 2, 50)
                    sheet2.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_uso_herramientas.xlsx"'
        return response

    # Exportar a PDF
    if "pdf" in request.GET:
        if rol == "mantenimiento":
            return HttpResponseForbidden("Solo puede consultar disponibilidad, no exportar reportes.")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=36, leftMargin=36, rightMargin=36, bottomMargin=36
        )
        styles = getSampleStyleSheet()
        story = []
        
        # Membrete
        _add_membrete_to_story(story, width=120, height=60)
        
        # Título
        story.append(Paragraph("Reporte de uso y disponibilidad de herramientas", styles["Title"]))
        story.append(Paragraph(
            f"Número total de herramientas registradas: {total_herramientas}", styles["Normal"]
        ))
        story.append(Spacer(1, 20))
        
        # Tabla principal - ANCHOS AJUSTADOS PARA TÍTULOS COMPLETOS
        headers = ["Herramienta", "Estado", "Tareas totales", "Riego", "Poda", "Fumigación"]
        data = [headers]
        for f in filas:
            data.append([
                f["nombre"], f["estado"], f["total"],
                f["riego"], f["poda"], f["fumigacion"],
            ])
        
        # ANCHOS OPTIMIZADOS: equilibrados para todos los campos
        tabla = Table(data, repeatRows=1, colWidths=[85, 50, 55, 45, 45, 55])
        tabla.setStyle(_get_default_table_style())
        story.append(tabla)
        story.append(Spacer(1, 20))
        
        # Historial (si hay)
        if historial:
            story.append(PageBreak())
            story.append(Paragraph("Historial de tareas vinculadas", styles["Heading2"]))
            story.append(Spacer(1, 12))
            
            data_hist = [["Herramienta", "Tipo tarea", "Planta", "Fecha", "Estado", "Usuario"]]
            for h in historial[:300]:
                data_hist.append([
                    h["herramienta"], h["tipo"], h["planta"],
                    h["fecha_programada"].strftime("%Y-%m-%d %H:%M") if h["fecha_programada"] else "",
                    h["estado"], h["usuario"],
                ])
            
            # ANCHOS PARA HISTORIAL
            tabla_hist = Table(data_hist, repeatRows=1, colWidths=[60, 60, 60, 70, 50, 80])
            tabla_hist.setStyle(_get_default_table_style(header_color="#4a5c6a"))
            story.append(tabla_hist)
        
        doc.build(story)
        pdf_value = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_uso_herramientas.pdf"'
        return response

    # Render HTML
    context = {
        "rol": rol, "total_herramientas": total_herramientas, "herramientas": filas,
        "historial": historial, "q": q, "estado_filtro": estado_filtro,
        "f_ini": request.GET.get("f_ini", ""), "f_fin": request.GET.get("f_fin", ""),
    }
    return render(request, "reportes/reportes_uso_herramientas.html", context)


# --------- 8) REPORTE DE PRODUCTOS ---------
@login_required
def reporte_productos_inventario(request):
    usuario = request.user
    rol = _rol(usuario)
    if rol not in ("administrador", "gestor", "mantenimiento"):
        return HttpResponseForbidden("No tiene permiso para ver este reporte.")

    # Filtros
    q = request.GET.get("q", "").strip()
    bajo_stock = request.GET.get("bajo_stock") == "1"
    f_ini = parse_date(request.GET.get("f_ini")) if request.GET.get("f_ini") else None
    f_fin = parse_date(request.GET.get("f_fin")) if request.GET.get("f_fin") else None

    # Query de productos
    qs = Producto.objects.select_related("existencia").all()
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))
    if f_ini:
        qs = qs.filter(fecha_llegada__gte=f_ini)
    if f_fin:
        qs = qs.filter(fecha_llegada__lte=f_fin)

    # Procesar datos
    hoy = timezone.now().date()
    productos = []
    for p in qs:
        existencias = p.existencias_actuales
        dias_inv = None
        if p.fecha_llegada:
            dias_inv = (hoy - p.fecha_llegada).days
        productos.append({
            "id": p.id_producto, "nombre": p.nombre, "descripcion": p.descripcion,
            "existencias": existencias, "fecha_llegada": p.fecha_llegada,
            "dias_inventario": dias_inv,
        })
    
    # Filtrar bajo stock
    if bajo_stock:
        productos = [x for x in productos if (x["existencias"] is not None and x["existencias"] < 10)]
    
    total_productos = len(productos)

    # Movimientos
    movimientos = []
    producto_ids = [p["id"] for p in productos]
    mov_qs = AsignacionProducto.objects.filter(producto_id__in=producto_ids)
    if f_ini:
        mov_qs = mov_qs.filter(fecha_asignacion__date__gte=f_ini)
    if f_fin:
        mov_qs = mov_qs.filter(fecha_asignacion__date__lte=f_fin)
    
    for m in mov_qs.select_related("producto", "asignado_por").order_by("-fecha_asignacion")[:1000]:
        movimientos.append({
            "producto": m.producto.nombre if m.producto else "",
            "cantidad": m.cantidad, "tipo": "Salida",
            "fecha": m.fecha_asignacion, "tarea_id": m.tarea_id,
            "asignado_por": (
                m.asignado_por.nombre_completo
                if m.asignado_por and getattr(m.asignado_por, "nombre_completo", None)
                else (m.asignado_por.matricula if m.asignado_por else "")
            ),
        })

    # Exportar a Excel
    if "excel" in request.GET:
        if rol == "mantenimiento":
            return HttpResponseForbidden("Solo puede consultar existencias, no exportar reportes.")
        
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        
        df_prod = pd.DataFrame(productos)
        df_mov = pd.DataFrame(movimientos)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja 1: Inventario
            df_prod.to_excel(writer, sheet_name="Inventario", index=False, startrow=6)
            workbook = writer.book
            sheet = writer.sheets["Inventario"]
            
            # Membrete
            logo_path = _get_membrete_path()
            if logo_path:
                logo = XLImage(logo_path)
                logo.width = 180
                logo.height = 90
                sheet.add_image(logo, "A1")
            
            # Encabezado
            sheet["A5"] = "Reporte de productos e inventario"
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)
            sheet["A6"] = (
                f"Total de productos: {total_productos} "
                f"| Productos con bajo stock (<10): "
                f"{len([p for p in productos if p['existencias'] < 10])}"
            )
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df_prod.columns, 1):
                max_length = max(
                    df_prod[col].astype(str).apply(len).max() if not df_prod.empty else 0,
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
            
            # Hoja 2: Movimientos
            if not df_mov.empty:
                df_mov.to_excel(writer, sheet_name="Movimientos", index=False, startrow=4)
                sheet2 = writer.sheets["Movimientos"]
                sheet2["A1"] = "Movimientos de existencias (salidas a tareas)"
                sheet2["A1"].font = sheet2["A1"].font.copy(bold=True, size=14)
                
                # Ajustar columnas del historial
                for idx, col in enumerate(df_mov.columns, 1):
                    max_length = max(
                        df_mov[col].astype(str).apply(len).max() if not df_mov.empty else 0,
                        len(col)
                    )
                    adjusted_width = min(max_length + 2, 50)
                    sheet2.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="reporte_productos_inventario.xlsx"'
        )
        return response

    # Exportar a PDF
    if "pdf" in request.GET:
        if rol == "mantenimiento":
            return HttpResponseForbidden("Solo puede consultar existencias, no exportar reportes.")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4, topMargin=36, leftMargin=36, rightMargin=36, bottomMargin=36
        )
        styles = getSampleStyleSheet()
        story = []
        
        # Membrete
        _add_membrete_to_story(story, width=120, height=60)
        
        # Título y estadísticas
        story.append(Paragraph("Reporte de productos e inventario", styles["Title"]))
        story.append(Paragraph(
            f"Total de productos: {total_productos} | Productos con bajo stock (&lt;10): "
            f"{len([p for p in productos if p['existencias'] < 10])}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 10))
        
        # Tabla de productos - ANCHOS AJUSTADOS PARA TÍTULOS COMPLETOS
        headers = ["Producto", "Descripción", "Existencias", "Fecha llegada", "Días en inventario"]
        data = [headers]
        for p in productos[:300]:
            data.append([
                p["nombre"], p["descripcion"], p["existencias"],
                p["fecha_llegada"].strftime("%Y-%m-%d") if p["fecha_llegada"] else "",
                p["dias_inventario"] if p["dias_inventario"] is not None else "",
            ])
        
        # ANCHOS OPTIMIZADOS: más espacio para "Descripción" y "Producto"
        tabla = Table(data, repeatRows=1, colWidths=[70, 160, 50, 55, 55])
        tabla.setStyle(_get_default_table_style())
        story.append(tabla)
        story.append(Spacer(1, 12))
        
        # Movimientos (si hay)
        if movimientos:
            story.append(Paragraph("Movimientos de existencias (salidas a tareas)", styles["Heading3"]))
            data_mov = [["Producto", "Cantidad", "Tipo", "Fecha", "Tarea", "Asignado por"]]
            for m in movimientos[:300]:
                data_mov.append([
                    m["producto"], m["cantidad"], m["tipo"],
                    m["fecha"].strftime("%Y-%m-%d %H:%M") if m["fecha"] else "",
                    m["tarea_id"], m["asignado_por"],
                ])
            
            # ANCHOS PARA MOVIMIENTOS
            tabla_mov = Table(data_mov, repeatRows=1, colWidths=[70, 45, 45, 60, 45, 70])
            tabla_mov.setStyle(_get_default_table_style(header_color="#4a5c6a"))
            story.append(tabla_mov)
        
        doc.build(story)
        pdf_value = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="reporte_productos_inventario.pdf"'
        )
        return response

    # Render HTML
    context = {
        "rol": rol, "total_productos": total_productos, "productos": productos,
        "movimientos": movimientos if rol in ("administrador", "gestor") else [],
        "q": q, "bajo_stock": bajo_stock,
        "f_ini": request.GET.get("f_ini", ""), "f_fin": request.GET.get("f_fin", ""),
    }
    return render(request, "reportes/reportes_productos_inventario.html", context)
