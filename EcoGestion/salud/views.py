from datetime import datetime
import os
from collections import Counter

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import get_user_model
from notificaciones.signals import notificar
from django.db.models import Q
from django.http import HttpResponse
from django.utils.timezone import localtime
from django.contrib.staticfiles import finders
from django.conf import settings

import pandas as pd
from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Image,
    Spacer,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

# Gráficas en PDF
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

from .models import SaludRegistro, SaludHistorial
from .forms import SaludForm


# =========================
#  HELPER DE ROLES / FECHAS
# =========================

def es_admin_o_gestor(user):
    return user.is_authenticated and getattr(user, "rol", None) in ["administrador", "gestor"]


def es_admin_gestor_mantenimiento(user):
    return user.is_authenticated and getattr(user, "rol", None) in ["administrador", "gestor", "mantenimiento"]


def _parse_date(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


# =========================
#  HELPERS DE DISEÑO (Membrete + Tablas + Gráficas)
# =========================

def _get_membrete_path():
    """
    Devuelve la ruta absoluta del logo/membrete de UPEMOR.
    Intenta primero en static/img/upemor-logo.png y luego con finders.
    """
    path_absoluto = os.path.join(settings.BASE_DIR, "static", "img", "upemor-logo.png")
    if os.path.exists(path_absoluto):
        return path_absoluto

    try:
        path_alternativo = finders.find("img/upemor-logo.png")
        if path_alternativo:
            return path_alternativo
    except Exception:
        pass

    print("ADVERTENCIA: No se encontró el archivo de membrete.")
    return None


def _add_membrete_to_story(story, width=120, height=60):
    """
    Inserta el logo en el PDF (si existe) y deja un pequeño espacio.
    No truena si el archivo no se encuentra.
    """
    membrete_path = _get_membrete_path()
    if membrete_path:
        try:
            img = Image(membrete_path, width=width, height=height)
            story.append(img)
            story.append(Spacer(1, 12))
        except Exception as e:
            print(f"Error al cargar membrete: {e}")
            story.append(Spacer(1, 20))
    else:
        story.append(Spacer(1, 20))


def _get_default_table_style(header_color="#6a67c8"):
    """
    Estilo de tabla reutilizable para PDFs.
    Cabecera coloreada, zebra rows, fuentes pequeñas y padding.
    """
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.whitesmoke, colors.HexColor("#f4f4f8")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]
    )


def _create_bar_chart(title_text, data_list, labels_list):
    """
    Crea un gráfico de barras con título integrado.
    data_list: lista de valores numéricos.
    labels_list: lista de categorías (strings).
    """
    drawing = Drawing(450, 220)

    # Título
    title = String(
        225,
        190,
        title_text,
        textAnchor="middle",
        fontSize=11,
        fontName="Helvetica-Bold",
    )
    drawing.add(title)

    if not data_list:
        return drawing

    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 50
    bc.height = 130
    bc.width = 380

    bc.data = [data_list]
    bc.categoryAxis.categoryNames = labels_list
    bc.categoryAxis.labels.angle = 0
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.dy = -10

    bc.valueAxis.valueMin = 0
    bc.valueAxis.labels.fontSize = 7
    bc.barWidth = 8
    bc.groupSpacing = 3

    drawing.add(bc)
    return drawing


def _create_pie_chart(title_text, data_list, labels_list):
    """
    Crea un gráfico de pastel con título integrado.
    Usa pie.slices.fontSize (correcto) en lugar de pie.labels.fontSize.
    """
    drawing = Drawing(450, 220)

    title = String(
        225,
        190,
        title_text,
        textAnchor="middle",
        fontSize=11,
        fontName="Helvetica-Bold",
    )
    drawing.add(title)

    if not data_list or sum(data_list) == 0:
        return drawing

    pie = Pie()
    pie.x = 150
    pie.y = 30
    pie.width = 150
    pie.height = 150
    pie.data = data_list
    pie.labels = labels_list
    pie.slices.fontSize = 7           # <- aquí está la corrección
    pie.slices.strokeWidth = 0.5
    pie.slices.strokeColor = colors.black

    drawing.add(pie)
    return drawing


# =========================
#  CRUD DE SALUD GENERAL
# =========================

@login_required
@user_passes_test(es_admin_gestor_mantenimiento)
def inicio(request):
    qs = SaludRegistro.objects.select_related('planta', 'usuario')

    q = (request.GET.get('q') or '').strip()
    estado = (request.GET.get('estado') or '').strip()
    f_from = request.GET.get('desde') or ''
    f_to = request.GET.get('hasta') or ''
    planta_id = request.GET.get('planta') or ''
    usuario_id = request.GET.get('usuario') or ''

    if q:
        qs = qs.filter(
            Q(planta__nombre_comun__icontains=q) |
            Q(planta__nombre_cientifico__icontains=q)
        )
    if estado in {'verde', 'amarillo', 'rojo'}:
        qs = qs.filter(estado_salud=estado)
    d_from = _parse_date(f_from)
    d_to = _parse_date(f_to)
    if d_from:
        qs = qs.filter(fecha_actualizacion__date__gte=d_from)
    if d_to:
        qs = qs.filter(fecha_actualizacion__date__lte=d_to)
    if planta_id.isdigit():
        qs = qs.filter(planta_id=int(planta_id))
    if usuario_id.isdigit():
        qs = qs.filter(usuario_id=int(usuario_id))

    registros = qs.order_by('-fecha_actualizacion', '-id_registro')

    ctx = {
        'registros': registros[:300],
        'q': q, 'estado': estado, 'desde': f_from, 'hasta': f_to,
        'planta_id': planta_id, 'usuario_id': usuario_id,
        'total': qs.count(),
    }
    return render(request, 'salud/read.html', ctx)


@login_required
@user_passes_test(es_admin_o_gestor)
def crear(request):
    if request.method == 'POST':
        form = SaludForm(request.POST)
        if form.is_valid():
            planta = form.cleaned_data['planta']
            if SaludRegistro.objects.filter(planta=planta).exists():
                return redirect('salud_inicio')

            obj = form.save(commit=False)
            obj.usuario = request.user
            obj.save()

            SaludHistorial.objects.create(
                planta=obj.planta, usuario=request.user,
                estado_salud=obj.estado_salud, observaciones=obj.observaciones,
            )
            _notificar_si_riesgo(obj)
            return redirect('salud_inicio')
    else:
        form = SaludForm()
    return render(request, 'salud/create.html', {'form': form})


@login_required
@user_passes_test(es_admin_o_gestor)
def editar(request, id_registro):
    reg = get_object_or_404(SaludRegistro, pk=id_registro)
    if request.method == 'POST':
        form = SaludForm(request.POST, instance=reg)
        form.fields['planta'].required = False
        form.fields['planta'].disabled = True

        if form.is_valid():
            obj = form.save(commit=False)
            obj.planta = reg.planta
            obj.usuario = request.user
            obj.save()

            # registrar historial cada vez que se edita
            SaludHistorial.objects.create(
                planta=obj.planta,
                usuario=request.user,
                estado_salud=obj.estado_salud,
                observaciones=obj.observaciones,
            )

            _notificar_si_riesgo(obj)
            return redirect('salud_inicio')
    else:
        form = SaludForm(instance=reg)
        form.fields['planta'].disabled = True

    return render(request, 'salud/update.html', {'form': form, 'registro': reg})


@login_required
@user_passes_test(es_admin_o_gestor)
def eliminar(request, id_registro):
    registro = get_object_or_404(SaludRegistro, pk=id_registro)

    if request.method == 'POST':
        registro.delete()
        return redirect('salud_inicio')

    return redirect('salud_inicio')


def _notificar_si_riesgo(reg: SaludRegistro):
    # solo notificamos si es amarillo o rojo
    if reg.estado_salud not in ('amarillo', 'rojo'):
        return

    User = get_user_model()
    admins = User.objects.filter(rol='administrador', is_active=True)

    if not admins.exists():
        return

    mensaje = (
        f"La planta '{reg.planta}' fue registrada con estado '{reg.estado_salud}'. "
        "Requiere atención prioritaria."
    )

    nivel = 'Atencion' if reg.estado_salud == 'amarillo' else 'Peligro'

    for admin in admins:
        notificar.send(
            sender=reg.usuario,
            destiny=admin,
            verbo=mensaje,
            target=reg.planta,
            level=nivel,
        )


# =========================
#  HISTORIAL POR PLANTA + EXCEL/PDF + 2 GRÁFICAS
# =========================

@login_required
@user_passes_test(es_admin_gestor_mantenimiento)
def historial_planta(request, planta_id):
    from plantas.models import plantaArbol
    planta = get_object_or_404(plantaArbol, pk=planta_id)

    # Query base del historial de esa planta
    qs = (
        SaludHistorial.objects
        .select_related('usuario')
        .filter(planta_id=planta_id)
        .order_by('-fecha_evento')
    )

    total = qs.count()

    # ----- Datos normalizados para tabla / exportación -----
    filas = []
    for h in qs:
        filas.append({
            "fecha": localtime(h.fecha_evento).strftime("%Y-%m-%d %H:%M"),
            "estado": (
                h.get_estado_salud_display()
                if hasattr(h, "get_estado_salud_display")
                else h.estado_salud
            ),
            "estado_codigo": h.estado_salud,
            "usuario": h.usuario.nombre_completo if h.usuario else "",
            "observaciones": h.observaciones or "",
        })

    # ----- Datos para gráficas (2 gráficas) -----
    # Gráfica 1: conteo por estado de salud
    contador_estados = Counter(f["estado_codigo"] for f in filas)
    estado_labels = []
    estado_data = []
    mapa_display = {"verde": "Verde", "amarillo": "Amarillo", "rojo": "Rojo"}

    for cod in ["verde", "amarillo", "rojo"]:
        if contador_estados.get(cod, 0) > 0:
            estado_labels.append(mapa_display.get(cod, cod.title()))
            estado_data.append(contador_estados[cod])

    # Gráfica 2: distribución de estados en porcentaje
    total_estados = sum(estado_data) if estado_data else 0
    estado_pct = [
        round((v * 100.0) / total_estados, 2) if total_estados > 0 else 0
        for v in estado_data
    ]
    labels_pct = [
        f"{lbl} ({val}%)" for lbl, val in zip(estado_labels, estado_pct)
    ]

    # =========================
    # 1) Exportar a EXCEL
    # =========================
    if 'excel' in request.GET:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter

        df = pd.DataFrame([{
            "Fecha": f["fecha"],
            "Estado": f["estado"],
            "Usuario": f["usuario"],
            "Observaciones": f["observaciones"],
        } for f in filas])

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja 1: Historial
            df.to_excel(
                writer,
                sheet_name="HistorialPlanta",
                index=False,
                startrow=7,  # espacio para logo y títulos
            )
            workbook = writer.book
            sheet = writer.sheets["HistorialPlanta"]

            # Logo
            logo_path = _get_membrete_path()
            if logo_path:
                try:
                    logo = XLImage(logo_path)
                    logo.width = 180
                    logo.height = 90
                    sheet.add_image(logo, "A1")
                except Exception:
                    pass

            # Título y resumen
            titulo = f"Historial de salud de {planta.nombre_comun} ({planta.nombre_cientifico})"
            sheet["A5"] = titulo
            sheet["A5"].font = sheet["A5"].font.copy(bold=True, size=14)

            sheet["A6"] = f"Registros totales: {total}"
            sheet["A6"].font = sheet["A6"].font.copy(bold=True)

            # Ajuste de ancho de columnas
            if not df.empty:
                for idx, col in enumerate(df.columns, 1):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

            # Hoja 2: Gráficas
            sheet2 = workbook.create_sheet(title="Gráficas")

            # Datos para gráfica 1: conteo por estado
            sheet2["A1"] = "Estado"
            sheet2["B1"] = "Cantidad"
            for i, (lbl, val) in enumerate(zip(estado_labels, estado_data), start=2):
                sheet2[f"A{i}"] = lbl
                sheet2[f"B{i}"] = int(val)

            # Gráfica 1: BarChart (conteo por estado)
            if estado_labels:
                bar = BarChart()
                bar.title = "Cantidad de registros por estado"
                bar.y_axis.title = "Registros"
                bar.x_axis.title = "Estado"

                data_ref = Reference(sheet2, min_col=2, min_row=1, max_row=1 + len(estado_data))
                cats_ref = Reference(sheet2, min_col=1, min_row=2, max_row=1 + len(estado_data))
                bar.add_data(data_ref, titles_from_data=True)
                bar.set_categories(cats_ref)
                bar.width = 20
                bar.height = 12
                sheet2.add_chart(bar, "A10")

            # Datos para gráfica 2: porcentajes por estado
            sheet2["D1"] = "Estado"
            sheet2["E1"] = "% del total"
            for i, (lbl, val) in enumerate(zip(estado_labels, estado_pct), start=2):
                sheet2[f"D{i}"] = lbl
                sheet2[f"E{i}"] = float(val)

            # Gráfica 2: PieChart (porcentaje por estado)
            if estado_labels:
                pie_chart = PieChart()
                pie_chart.title = "Distribución porcentual por estado"

                data_ref2 = Reference(sheet2, min_col=5, min_row=1, max_row=1 + len(estado_pct))
                cats_ref2 = Reference(sheet2, min_col=4, min_row=2, max_row=1 + len(estado_pct))
                pie_chart.add_data(data_ref2, titles_from_data=True)
                pie_chart.set_categories(cats_ref2)
                pie_chart.width = 20
                pie_chart.height = 12
                sheet2.add_chart(pie_chart, "J10")

        filename = f"historial_salud_{planta.nombre_comun}.xlsx".replace(" ", "_")
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return response

    # =========================
    # 2) Exportar a PDF
    # =========================
    if 'pdf' in request.GET:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=36,
            leftMargin=36,
            rightMargin=36,
            bottomMargin=36,
        )
        styles = getSampleStyleSheet()
        story = []

        # Logo / membrete
        _add_membrete_to_story(story, width=120, height=60)

        # Título
        story.append(Paragraph(
            f"Historial de salud de {planta.nombre_comun} ({planta.nombre_cientifico})",
            styles["Title"]
        ))
        story.append(Paragraph(f"Registros totales: {total}", styles["Normal"]))
        story.append(Spacer(1, 10))

        # Gráfica 1: barra (conteo por estado)
        if estado_labels:
            chart1 = _create_bar_chart(
                "Cantidad de registros por estado",
                estado_data,
                estado_labels,
            )
            story.append(chart1)
            story.append(Spacer(1, 15))

        # Gráfica 2: pie (porcentaje por estado)
        if estado_labels:
            chart2 = _create_pie_chart(
                "Distribución porcentual por estado",
                estado_pct,
                labels_pct,   # etiquetas con porcentaje
            )
            story.append(chart2)
            story.append(Spacer(1, 15))

        # Tabla de datos
        headers = ["Fecha", "Estado", "Usuario", "Observaciones"]
        data = [headers]
        for f in filas[:500]:
            data.append([
                f["fecha"],
                f["estado"],
                f["usuario"],
                f["observaciones"],
            ])

        # Anchos razonables para que quepa todo
        table = Table(
            data,
            repeatRows=1,
            colWidths=[70, 55, 80, 260],
        )
        table.setStyle(_get_default_table_style(header_color="#39a96f"))
        story.append(table)

        doc.build(story)

        pdf_value = buffer.getvalue()
        buffer.close()

        filename = f"historial_salud_{planta.nombre_comun}.pdf".replace(" ", "_")
        response = HttpResponse(pdf_value, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return response

    # =========================
    # 3) Render normal HTML
    # =========================
    ctx = {
        'planta': planta,
        'eventos': qs,
        'total': total,
    }
    return render(request, 'salud/historial_planta.html', ctx)
